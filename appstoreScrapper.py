from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
import datetime

'''
Este script se conecta a la pagina web de Google Apps Store y extrae una lista de las URLs
de cada una de las aplicaciones en el App Store
'''

base_url = ("https://play.google.com/store/apps")
driver = webdriver.Firefox(executable_path = 'C:\\WebDrivers\\geckodriver.exe')

url = (base_url) 
driver.get(url)
time.sleep(5)


lista = [ ]
#Realiza una lista de todas las url de la appStore
elems = driver.find_elements_by_xpath("//a[@href]")
for elem in elems:
    urls = elem.get_attribute("href")
    lista.append(urls)

#abre un archivo excel utilizando la libreria openpyxl
book = Workbook()
sheet = book.active
workbook_name = 'AppStoreList.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

page.append(lista)

wb.save(filename=workbook_name)

