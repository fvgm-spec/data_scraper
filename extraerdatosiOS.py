from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
import datetime

'''
Este script se conecta a la pagina web de iOS Apps Store que el usuario haya seleccionado
y extrae una lista de las los datos correspondientes al nombre de usuario, comentario y 
fecha del comentario'''

driver = webdriver.Firefox(executable_path = 'C:\\WebDrivers\\geckodriver.exe')
base_url = input("Escribir la URL de la App: ")
url = (base_url)
driver.get(url)
time.sleep(20)

nombres = [ ]
names = driver.find_elements_by_xpath("//span[contains(@class,'we-truncate we-truncate--single-line ember-view we-customer-review__user')]")
for name in names:
    n = name.text
    nombres.append(n)

comentarios = [ ]
comments = driver.find_elements_by_xpath("//p[contains(@dir,'ltr')]")
for comment in comments:
    c = comment.text
    comentarios.append(c)

fechas = [ ]
dates = driver.find_elements_by_xpath("//time[contains(@class,'we-customer-review__date')]")
for date in dates:
    d = date.text
    fechas.append(d)

#Abre el archivo excel correspondiente a las aplicaciones extraidas
wb = Workbook()
wb_name = 'appsiOS.xlsx'
sheet = wb.active
sheet.title = "apps"
row = 1
col_names = 1 #columna donde se escriben los nombres de usuario
col_comm = 2 #columna donde se escriben los comentarios de usuario
col_dates = 3 #columna donde se escriben las fechas

for nombre, comentario, fecha in zip(nombres, comentarios, fechas):
    sheet.cell(column=col_names, row=row, value=nombre)
    sheet.cell(column=col_comm, row=row, value=comentario)
    sheet.cell(column=col_dates, row=row, value=fecha)
    row+=1

wb.save(filename=wb_name)
